import json
import os
import time
import smtplib
import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from bs4 import BeautifulSoup
from lxml import html
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime


# Define the email function
def send_emails(email_list):
    for person in email_list:
        # Make the body of the email
        body = \
            f"""Bonjour,\n
        Le marché EuroNextParis vient de cloturé.\n
        Veuillez trouver ci-joint le fichier contenant les données des entreprises du CAC40 pour le {datetime.now().strftime("%d/%m/%Y")}.\n
        Cordialement,\n

        Trading-Bot\n
        IntelliTrade"""
        # make a MIME object to define parts of the email
        msg = MIMEMultipart()
        msg['From'] = email_from
        msg['To'] = person
        msg['Subject'] = f"""INFO : Valeurs des actions du CAC40 du {datetime.now().strftime("%d/%m/%Y")} Groupe 02"""

        # Attach the body of the message
        msg.attach(MIMEText(body, 'plain'))

        # Definir la piece jointe
        filename = "data/output/donnees_boursorama_" + datetime.now().strftime("%Y%m%d") + ".xlsx"

        # Lire le fichier python comme binary
        attachment = open(filename, 'rb')  # r for read and b for binary

        # Encode sur une base 64
        attachment_package = MIMEBase('application', 'octet-stream')
        attachment_package.set_payload((attachment).read())
        encoders.encode_base64(attachment_package)
        attachment_package.add_header('Content-Disposition', "attachment; filename= " + filename)
        msg.attach(attachment_package)
        # Cast as string
        text = msg.as_string()

        # Connection avec le serveur
        print("Connecting to server...")
        SMTP_server = smtplib.SMTP(smtp_server, smtp_port)
        SMTP_server.starttls()
        SMTP_server.login(email_from, pswd)
        print("Succesfully connected to server")
        # Send emails to "person" as the list is iterated
        print(f"Sending email to: {person}...")
        SMTP_server.sendmail(email_from, person, text)
        print(f"Email sent to: {person}")
    # Close the port
    SMTP_server.quit()


def scrape_boursorama(company):
    Symbol = company["Symbol"].split(".")[0]
    url = f'https://www.boursorama.com/cours/1rP{Symbol}/'
    response = requests.get(url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        root = html.fromstring(response.content)

        # Récupération du nom de l'entreprise
        name = company["Name"]

        # Récupération du cours de l'entreprise
        element = root.xpath("/html/body/main/div/section/header/div/div/div[1]/div[1]/div/div[1]/span[1]")
        open_price = element[0].text_content().replace("  ", " ").strip() if element else None

        # Récupération HIGH
        xpath1 = "/html/body/main/div/section/header/div/div/div[3]/div[1]/div/ul/li[3]/p[2]/span"
        xpath2 = "/html/body/main/div/section/header/div/div/div[3]/div[1]/div/ul/li[4]/p[2]/span"
        xpath3 = "/html/body/main/div/section/header/div/div/div[3]/div[2]/div/ul/li[1]/p[2]/span"
        xpath4 = "/html/body/main/div/section/header/div/div/div[3]/div[1]/div/ul/li[2]/p[2]/span"
        high = root.xpath(xpath1)
        low = root.xpath(xpath2)
        volume = root.xpath(xpath3)
        close = root.xpath(xpath4)

        high = high[0].text_content().replace("  ", " ").strip() if high else None
        low = low[0].text_content().replace("  ", " ").strip() if low else None
        volume = volume[0].text_content().replace("  ", " ").strip() if volume else None
        close = close[0].text_content().replace("  ", " ").strip() if close else None

        # Obtention de la date et de l'heure actuelles
        now = datetime.now()
        date_time = now.strftime("%Y-%m-%d %H:%M:%S")

        return {"Name": name, "Open": open_price, "Date et Heure": date_time, "High": high, "Low": low, "Close": close,
                "Volume": volume}
    else:
        print(f"Échec de la requête pour {company['Name']}. Code d'état:", response.status_code)
        return None


def scrape_all_companies(companies):
    data = []
    for index, row in companies.iterrows():
        result = scrape_boursorama(row)
        if result:
            data.append(result)
    return data


def insert_data_into_database(data):
    try:
        # Remplacer "N/A" par None dans toutes les colonnes numériques
        for entry in data:
            for key, value in entry.items():
                if key not in ['Name', 'Date et Heure']:
                    entry[key] = None if pd.isna(value) or value == "N/A" else value

        df = pd.DataFrame(data)

        # Convertir les colonnes en double précision
        df['high'] = df['high'].str.replace(' ', '').astype(float)
        df['low'] = df['low'].str.replace(' ', '').astype(float)
        df['open'] = df['open'].str.replace(' ', '').astype(float)
        df['close'] = df['close'].str.replace(' ', '').astype(float)

        # Autres colonnes à convertir si nécessaire
        df['volume'] = df['volume'].str.replace(' ', '').astype(float)

        # Insérer les données dans la table PostgreSQL
        df.to_sql(nom_table, con=engine, index=False, if_exists='append', method='multi')
        print(f"Données ajoutées avec succès à la table '{nom_table}' dans la base de données.")
    except Exception as e:
        print(f"Erreur lors de l'insertion des données : {e}")


def save_to_excel(data):
    try:
        filename = f'data/output/donnees_boursorama_{datetime.now().strftime("%Y%m%d")}.xlsx'  # Nom du fichier avec la date du jour
        # Vérifier si le fichier existe déjà
        if os.path.exists(filename):
            # Charger le fichier existant
            wb = load_workbook(filename)
            ws = wb.active
        else:
            # Créer un nouveau fichier Excel s'il n'existe pas
            wb = Workbook()
            ws = wb.active
            # Ajouter les en-têtes si le fichier est nouveau
            headers = list(data[0].keys())
            ws.append(headers)

        # Écrire les données dans le fichier Excel
        for row_data in data:
            row_values = list(row_data.values())
            ws.append(row_values)

        # Sauvegarder le fichier Excel
        wb.save(filename)
        print(f"Données ajoutées avec succès à {filename}.")
    except Exception as e:
        print(f"Erreur lors de l'enregistrement des données dans le fichier Excel : {e}")


def job():
    # Ajouter une vérification du jour de la semaine
    current_day = datetime.now().weekday()
    # Si le jour de la semaine est compris entre lundi (0) et vendredi (4)
    if 0 <= current_day <= 4:
        print("Exécution de la tâche...")
        all_data = scrape_all_companies(cac40_companies)
        insert_data_into_database(all_data)
        save_to_excel(all_data)
    else:
        print("Le programme ne s'exécute pas le week-end.")


# Charger la liste des entreprises depuis le fichier Excel
company_referal_path = 'data/input/liste_cac40.xlsx'
cac40_companies = pd.read_excel(company_referal_path)

# Informations de connexion à la base de données Alwaysdata
with open(os.path.expandvars("config/bd_config.json")) as json_file:
    conn_params = json.load(json_file)

# Nom de la table dans la base de données
nom_table = 'data'

# Créer une connexion à la base de données avec SQLAlchemy
engine = create_engine(
    f'postgresql+psycopg2://{conn_params["user"]}:{conn_params["password"]}@{conn_params["host"]}:{conn_params["port"]}/{conn_params["database"]}')

# Planifier l'exécution du travail chaque jour de 9h à 18h
start_time = datetime.now().replace(hour=8, minute=30, second=0, microsecond=0)
end_time = datetime.now().replace(hour=18, minute=0, second=0, microsecond=0)

while datetime.now() < end_time:
    if datetime.now() >= start_time:
        job()
    # Attendre 30 minutes avant de vérifier à nouveau
    time.sleep(1800)

print("Programme terminé)" + str(datetime.now()))
with open(os.path.expandvars("config/mail_config.json")) as json_file:
    gmail_cfg = json.load(json_file)

# Setup smtp server params
smtp_port = gmail_cfg["port"]  # Standard secure SMTP port
smtp_server = gmail_cfg["server"]  # Google SMTP Server

# Set up the email lists
email_from = gmail_cfg["sender"]
email_list = gmail_cfg["reciever"]

# Define the password
pswd = gmail_cfg["password"]

# Run the function
send_emails(email_list)