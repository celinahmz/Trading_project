
# VERSION 2.5 :     ----------------------  CODE QUI FONCTIONNE Execution temps reel code complet chaque trois minute ------------------------------

import os
from openpyxl import Workbook, load_workbook
import time
import pandas as pd
from datetime import datetime
from sqlalchemy import create_engine
import numpy as np
import requests
from bs4 import BeautifulSoup
from lxml import html

# Charger la liste des entreprises depuis le fichier Excel
company_referal_path = 'liste_cac40.xlsx'
cac40_companies = pd.read_excel(company_referal_path)

# Informations de connexion à la base de données Alwaysdata
conn_params = {
    'database': 'projetbourse_trading',
    'user': 'projetbourse_',
    'password': 'SSDC2024',
    'host': 'postgresql-projetbourse.alwaysdata.net',
    'port': '5432'
}

# Nom de la table dans la base de données
nom_table = 'data'

# Créer une connexion à la base de données avec SQLAlchemy
engine = create_engine(
    f'postgresql+psycopg2://{conn_params["user"]}:{conn_params["password"]}@{conn_params["host"]}:{conn_params["port"]}/{conn_params["database"]}')


def scrape_boursorama(company):
    Symbol = company["Symbol"].split(".")[0]
    url = f'https://www.boursorama.com/cours/{Symbol}/'
    response = requests.get(url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        root = html.fromstring(response.content)

        # Récupération du nom de l'entreprise
        name = company["Name"]

        # Récupération du cours de l'entreprise
        element = root.xpath("/html/body/main/div/section/header/div/div/div[1]/div[1]/div/div[1]/span[1]")
        open_price = element[0].text_content().replace("  ", " ").strip() if element else None

        # Récupération HIGH
        xpath1 = "/html/body/main/div/section/header/div/div/div[3]/div[1]/div/ul/li[3]/p[2]/span"
        xpath2 = "/html/body/main/div/section/header/div/div/div[3]/div[1]/div/ul/li[4]/p[2]/span"
        xpath3 = "/html/body/main/div/section/header/div/div/div[3]/div[2]/div/ul/li[1]/p[2]/span"

        element1 = root.xpath(xpath1)
        element2 = root.xpath(xpath2)
        element3 = root.xpath(xpath3)

        element1 = element1[0].text_content().replace("  ", " ").strip() if element1 else None
        element2 = element2[0].text_content().replace("  ", " ").strip() if element2 else None
        element3 = element3[0].text_content().replace("  ", " ").strip() if element3 else None

        # Obtention de la date et de l'heure actuelles
        now = datetime.now()
        date_time = now.strftime("%Y-%m-%d %H:%M:%S")

        return {"Name": name, "Open": open_price, "Date et Heure": date_time, "High": element1, "Low": element2,
                "Volume": element3}
    else:
        print(f"Échec de la requête pour {company['Name']}. Code d'état:", response.status_code)
        return None


def scrape_all_companies(companies):
    data = []
    for index, row in companies.iterrows():
        result = scrape_boursorama(row)
        if result:
            data.append(result)
    return data


def insert_data_into_database(data):
    try:
        # Remplacer "N/A" par None dans toutes les colonnes numériques
        for entry in data:
            for key, value in entry.items():
                if key not in ['Name', 'Date et Heure']:
                    entry[key] = None if pd.isna(value) or value == "N/A" else value

        df = pd.DataFrame(data)

        # Convertir les colonnes en double précision
        df['High'] = df['High'].str.replace(' ', '').astype(float)
        df['Low'] = df['Low'].str.replace(' ', '').astype(float)
        df['Open'] = df['Open'].str.replace(' ', '').astype(float)

        # Autres colonnes à convertir si nécessaire
        df['Volume'] = df['Volume'].str.replace(' ', '').astype(float)

        # Insérer les données dans la table PostgreSQL
        df.to_sql(nom_table, con=engine, index=False, if_exists='append', method='multi')
        print(f"Données ajoutées avec succès à la table '{nom_table}' dans la base de données.")
    except Exception as e:
        print(f"Erreur lors de l'insertion des données : {e}")


def save_to_excel(data):
    try:
        filename = f'donnees_boursorama_{datetime.now().strftime("%Y%m%d")}.xlsx'  # Nom du fichier avec la date du jour
        # Vérifier si le fichier existe déjà
        if os.path.exists(filename):
            # Charger le fichier existant
            wb = load_workbook(filename)
            ws = wb.active
        else:
            # Créer un nouveau fichier Excel s'il n'existe pas
            wb = Workbook()
            ws = wb.active
            # Ajouter les en-têtes si le fichier est nouveau
            headers = list(data[0].keys())
            ws.append(headers)

        # Écrire les données dans le fichier Excel
        for row_data in data:
            row_values = list(row_data.values())
            ws.append(row_values)

        # Sauvegarder le fichier Excel
        wb.save(filename)
        print(f"Données ajoutées avec succès à {filename}.")
    except Exception as e:
        print(f"Erreur lors de l'enregistrement des données dans le fichier Excel : {e}")


def job():
    print("Exécution de la tâche...")
    all_data = scrape_all_companies(cac40_companies)
    insert_data_into_database(all_data)
    save_to_excel(all_data)


# Planifier l'exécution du travail chaque jour de 9h à 18h
start_time = datetime.now().replace(hour=19, minute=39, second=0, microsecond=0)
end_time = datetime.now().replace(hour=19, minute=50, second=0, microsecond=0)
while datetime.now() < end_time:
    if datetime.now() >= start_time:
        job()
    time.sleep(180)  # Attendre 5 minutes avant de vérifier à nouveau

print("Programme terminé)" + str(datetime.now()))
